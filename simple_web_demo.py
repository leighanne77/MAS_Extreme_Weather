#!/usr/bin/env python3
"""
Pythia Mobile Bay Case Study Demo

A demonstration of the Pythia extreme weather risk analysis system
focused on the Mobile Bay, Alabama prototype for Private Equity Investors.

This demo uses REAL DATA from the JSON files to show:
- Gulf Coast regional risk profiles
- Nature-based solutions with actual case studies
- Economic impact analysis
- Funding sources for resilience projects

Run with: python simple_web_demo.py
Open: http://localhost:8000
"""

import sys
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

# Add the src directory to the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn
import json
import io

# PDF Export
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel Export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import our data source manager (replaces legacy DataLoader)
from multi_agent_system.data import get_data_source_manager
from multi_agent_system.risk_definitions import RiskLevel

# =============================================================================
# APP CONFIGURATION
# =============================================================================

app = FastAPI(
    title="Pythia - Mobile Bay Case Study Demo",
    description="Extreme Weather Risk Analysis for Private Equity Investors",
    version="1.0.0"
)

# Mount static files
static_path = Path(__file__).parent / "src" / "tool_web" / "static"
if static_path.exists():
    app.mount("/static", StaticFiles(directory=str(static_path)), name="static")

# Templates
templates_path = Path(__file__).parent / "src" / "tool_web" / "templates"
if templates_path.exists():
    templates = Jinja2Templates(directory=str(templates_path))
else:
    templates = None

# Data source manager - connects to real JSON data (replaces legacy data_loader)
data_loader = get_data_source_manager()

# =============================================================================
# CONSTANTS
# =============================================================================

# User types (7 types, PE Investor is primary prototype)
USER_TYPES = [
    {"id": "private_equity", "name": "Private Equity Investor", "primary": True},
    {"id": "private_debt", "name": "Private Debt Manager", "primary": False},
    {"id": "chief_risk_officer", "name": "Chief Risk Officer", "primary": False},
    {"id": "chief_sustainability_officer", "name": "Chief Sustainability Officer", "primary": False},
    {"id": "data_science_officer", "name": "Data Science Officer", "primary": False},
    {"id": "operating_credit_officer", "name": "Operating Credit Officer", "primary": False},
    {"id": "government_funder", "name": "Government Funder", "primary": False},
]

# Primary prototype location
PRIMARY_PROTOTYPE = {
    "name": "Mobile Bay, Alabama",
    "region": "gulf_coast",
    "coordinates": {"lat": 30.6954, "lon": -88.0399},
    "description": "Coastal manufacturing hub with significant hurricane and storm surge exposure"
}

# Decision support disclaimer
DISCLAIMER = """
ADVISORY ANALYSIS - REQUIRES PROFESSIONAL REVIEW
Pythia is a decision support tool, NOT a decision making tool. 
ROI analysis frameworks provided - no guarantees.
Export-based integration only.
"""

# =============================================================================
# DATA ACCESS HELPERS
# =============================================================================

def get_gulf_coast_profile() -> Dict[str, Any]:
    """Get the Gulf Coast regional risk profile from real data."""
    profiles = data_loader.load_regional_risk_profiles()
    return profiles.get("regions", {}).get("gulf_coast", {})


def get_mobile_bay_solutions() -> List[Dict[str, Any]]:
    """Get nature-based solutions relevant to Mobile Bay with case studies."""
    nbs_data = data_loader.load_nature_based_solutions()
    solutions = nbs_data.get("solutions", [])
    
    # Filter solutions relevant to coastal/flooding risks
    relevant_risk_types = {"flooding", "storm_surge", "hurricanes", "coastal_erosion", "water_quality"}
    
    mobile_bay_solutions = []
    for solution in solutions:
        solution_risks = set(solution.get("risk_types", []))
        if solution_risks & relevant_risk_types:
            # Check for Mobile Bay case studies
            case_studies = solution.get("case_studies", [])
            mobile_bay_cases = [
                cs for cs in case_studies 
                if "Mobile Bay" in cs.get("location", "") or "Alabama" in cs.get("location", "")
            ]
            solution_copy = solution.copy()
            solution_copy["mobile_bay_case_studies"] = mobile_bay_cases
            solution_copy["has_local_case_study"] = len(mobile_bay_cases) > 0
            mobile_bay_solutions.append(solution_copy)
    
    # Sort by relevance (local case studies first)
    mobile_bay_solutions.sort(key=lambda x: (not x["has_local_case_study"], x.get("name", "")))
    
    return mobile_bay_solutions


def get_gulf_coast_funding() -> List[Dict[str, Any]]:
    """Get funding sources applicable to Gulf Coast projects."""
    try:
        funding_path = Path(__file__).parent / "src" / "multi_agent_system" / "data" / "funding_sources_NSB.json"
        with open(funding_path, 'r') as f:
            funding_data = json.load(f)
        
        all_sources = []
        for category in funding_data.get("nature_based_solutions_funding", {}).get("funding_sources", []):
            for source in category.get("sources", []):
                source["category"] = category.get("category", "Unknown")
                all_sources.append(source)
        return all_sources
    except Exception:
        return []


def get_coastal_infrastructure_data() -> Dict[str, Any]:
    """Get coastal infrastructure needs data for Mobile Bay."""
    try:
        infra_path = Path(__file__).parent / "src" / "multi_agent_system" / "data" / "coastal_areas_infrastructure_needs.json"
        with open(infra_path, 'r') as f:
            return json.load(f)
    except Exception:
        return {}


def calculate_risk_score(profile: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate composite risk score from regional profile."""
    primary_risks = profile.get("primary_risks", {})
    
    risk_levels = {
        "extreme": 95,
        "high": 75,
        "medium": 50,
        "low": 25
    }
    
    scores = []
    risk_breakdown = []
    
    for risk_name, risk_data in primary_risks.items():
        level = risk_data.get("risk_level", "medium")
        score = risk_levels.get(level, 50)
        scores.append(score)
        risk_breakdown.append({
            "name": risk_name.replace("_", " ").title(),
            "level": level,
            "score": score,
            "details": risk_data
        })
    
    composite_score = sum(scores) / len(scores) if scores else 50
    
    return {
        "composite_score": round(composite_score, 1),
        "risk_level": "Extreme" if composite_score >= 85 else "High" if composite_score >= 65 else "Medium" if composite_score >= 40 else "Low",
        "confidence": 0.87,
        "breakdown": sorted(risk_breakdown, key=lambda x: x["score"], reverse=True)
    }


def calculate_economic_impact(hurricane_category: int = 3) -> Dict[str, Any]:
    """Calculate economic impact using real data."""
    economic_data = data_loader.load_economic_impact_data()
    hurricane_impacts = economic_data.get("economic_impacts", {}).get("hurricane_impacts", {})
    
    category_key = f"category_{hurricane_category}"
    impact = hurricane_impacts.get(category_key, {})
    
    return {
        "hurricane_category": hurricane_category,
        "damage_range": impact.get("damage_range", [0, 0]),
        "business_interruption": impact.get("business_interruption", [0, 0]),
        "insurance_claims": impact.get("insurance_claims", [0, 0]),
        "recovery_time_days": impact.get("recovery_time_days", [0, 0]),
        "formatted": {
            "damage": f"${impact.get('damage_range', [0, 0])[0]:,} - ${impact.get('damage_range', [0, 0])[1]:,}",
            "interruption": f"${impact.get('business_interruption', [0, 0])[0]:,} - ${impact.get('business_interruption', [0, 0])[1]:,}",
            "recovery": f"{impact.get('recovery_time_days', [0, 0])[0]} - {impact.get('recovery_time_days', [0, 0])[1]} days"
        }
    }


def get_historical_gulf_events() -> List[Dict[str, Any]]:
    """Get historical weather events for Gulf Coast region."""
    events_data = data_loader.load_historical_weather_events()
    events = events_data.get("events", [])
    
    gulf_events = [
        e for e in events 
        if "Gulf" in e.get("location", {}).get("region", "") or
        "Alabama" in str(e.get("location", {}).get("states", []))
    ]
    
    return gulf_events


# =============================================================================
# EXPORT HELPERS
# =============================================================================

def generate_pdf_report(analysis: Dict[str, Any]) -> io.BytesIO:
    """Generate a PDF report from analysis data."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, alignment=TA_CENTER, spaceAfter=30)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20)
    normal_style = styles['Normal']
    
    story = []
    
    # Title
    story.append(Paragraph("Pythia Climate Risk Analysis Report", title_style))
    story.append(Paragraph(f"Mobile Bay, Alabama - {datetime.now().strftime('%B %d, %Y')}", 
                          ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER)))
    story.append(Spacer(1, 30))
    
    # Disclaimer
    disclaimer_style = ParagraphStyle('Disclaimer', parent=styles['Normal'], fontSize=9, 
                                      backColor=colors.Color(1, 0.95, 0.85), leftIndent=10, rightIndent=10)
    story.append(Paragraph("<b>ADVISORY ANALYSIS - REQUIRES PROFESSIONAL REVIEW</b><br/>"
                          "Pythia is a decision support tool, NOT a decision making tool. "
                          "ROI analysis frameworks provided - no guarantees.", disclaimer_style))
    story.append(Spacer(1, 20))
    
    # Location
    story.append(Paragraph("Location Overview", heading_style))
    location = analysis.get("location", {})
    story.append(Paragraph(f"<b>Name:</b> {location.get('name', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Region:</b> {location.get('region', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Description:</b> {location.get('description', 'N/A')}", normal_style))
    story.append(Spacer(1, 10))
    
    # Risk Assessment
    story.append(Paragraph("Risk Assessment", heading_style))
    risk = analysis.get("risk_assessment", {})
    
    risk_data = [
        ["Risk Category", "Level", "Score"],
        ["Overall Risk", risk.get("risk_level", "N/A"), str(risk.get("risk_score", "N/A"))],
    ]
    
    risk_breakdown = risk.get("risk_breakdown", {})
    for category, details in risk_breakdown.items():
        if isinstance(details, dict):
            risk_data.append([category.replace("_", " ").title(), 
                            details.get("level", "N/A"), 
                            str(details.get("score", "N/A"))])
    
    risk_table = Table(risk_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.25)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 0.95)),
        ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.7, 0.7, 0.7)),
    ]))
    story.append(risk_table)
    story.append(Spacer(1, 20))
    
    # Nature-Based Solutions
    story.append(Paragraph("Recommended Nature-Based Solutions", heading_style))
    solutions = analysis.get("solutions", [])[:5]  # Top 5
    
    for i, sol in enumerate(solutions, 1):
        story.append(Paragraph(f"<b>{i}. {sol.get('name', 'N/A')}</b>", normal_style))
        story.append(Paragraph(f"   Relevance: {sol.get('relevance_score', 'N/A')}% | "
                              f"ROI: {sol.get('roi_range', 'N/A')} | "
                              f"Payback: {sol.get('payback_period', 'N/A')}", 
                              ParagraphStyle('Small', parent=styles['Normal'], fontSize=9)))
        if sol.get("has_local_case_study"):
            story.append(Paragraph("   ✓ Has Mobile Bay case study", 
                                  ParagraphStyle('CaseStudy', parent=styles['Normal'], fontSize=9, textColor=colors.darkgreen)))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 20))
    
    # Economic Impact
    story.append(Paragraph("Economic Impact Summary", heading_style))
    impact = analysis.get("economic_impact", {})
    story.append(Paragraph(f"<b>Scenario:</b> {impact.get('scenario', 'N/A')}", normal_style))
    formatted = impact.get("formatted", {})
    story.append(Paragraph(f"<b>Property Damage:</b> {formatted.get('damage', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Business Interruption:</b> {formatted.get('interruption', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Recovery Time:</b> {formatted.get('recovery', 'N/A')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Funding Sources
    story.append(Paragraph("Available Funding Sources", heading_style))
    funding = analysis.get("funding_sources", [])[:5]
    
    funding_data = [["Source", "Type", "Amount Range"]]
    for fund in funding:
        funding_data.append([
            fund.get("name", "N/A")[:40],
            fund.get("type", "N/A"),
            fund.get("amount_range", "N/A")
        ])
    
    funding_table = Table(funding_data, colWidths=[3*inch, 1.2*inch, 1.5*inch])
    funding_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.3, 0.5)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.95, 0.98)),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
    ]))
    story.append(funding_table)
    story.append(Spacer(1, 30))
    
    # Footer
    story.append(Paragraph(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                          ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)))
    story.append(Paragraph("Pythia Extreme Weather Risk Analysis System", 
                          ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_excel_report(analysis: Dict[str, Any]) -> io.BytesIO:
    """Generate an Excel report from analysis data."""
    buffer = io.BytesIO()
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "Pythia Climate Risk Analysis Report"
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Mobile Bay, Alabama - {datetime.now().strftime('%B %d, %Y')}"
    ws.merge_cells('A2:D2')
    
    ws['A4'] = "LOCATION"
    ws['A4'].font = title_font
    location = analysis.get("location", {})
    ws['A5'] = "Name:"
    ws['B5'] = location.get("name", "N/A")
    ws['A6'] = "Region:"
    ws['B6'] = location.get("region", "N/A")
    ws['A7'] = "Description:"
    ws['B7'] = location.get("description", "N/A")
    
    ws['A9'] = "RISK ASSESSMENT"
    ws['A9'].font = title_font
    risk = analysis.get("risk_assessment", {})
    ws['A10'] = "Overall Risk Level:"
    ws['B10'] = risk.get("risk_level", "N/A")
    ws['A11'] = "Risk Score:"
    ws['B11'] = risk.get("risk_score", "N/A")
    
    # Sheet 2: Risk Breakdown
    ws2 = wb.create_sheet("Risk Breakdown")
    headers = ["Category", "Level", "Score", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    risk_breakdown = risk.get("risk_breakdown", {})
    row = 2
    for category, details in risk_breakdown.items():
        if isinstance(details, dict):
            ws2.cell(row=row, column=1, value=category.replace("_", " ").title()).border = border
            ws2.cell(row=row, column=2, value=details.get("level", "N/A")).border = border
            ws2.cell(row=row, column=3, value=details.get("score", "N/A")).border = border
            ws2.cell(row=row, column=4, value=details.get("description", "")).border = border
            row += 1
    
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 3: Solutions
    ws3 = wb.create_sheet("Solutions")
    sol_headers = ["Solution", "Relevance %", "ROI Range", "Payback Period", "Has Case Study"]
    for col, header in enumerate(sol_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    solutions = analysis.get("solutions", [])
    for row, sol in enumerate(solutions, 2):
        ws3.cell(row=row, column=1, value=sol.get("name", "N/A")).border = border
        ws3.cell(row=row, column=2, value=sol.get("relevance_score", "N/A")).border = border
        ws3.cell(row=row, column=3, value=sol.get("roi_range", "N/A")).border = border
        ws3.cell(row=row, column=4, value=sol.get("payback_period", "N/A")).border = border
        ws3.cell(row=row, column=5, value="Yes" if sol.get("has_local_case_study") else "No").border = border
    
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 14
    
    # Sheet 4: Funding Sources
    ws4 = wb.create_sheet("Funding Sources")
    fund_headers = ["Source Name", "Type", "Amount Range", "Eligibility"]
    for col, header in enumerate(fund_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.border = border
    
    funding = analysis.get("funding_sources", [])
    for row, fund in enumerate(funding, 2):
        ws4.cell(row=row, column=1, value=fund.get("name", "N/A")).border = border
        ws4.cell(row=row, column=2, value=fund.get("type", "N/A")).border = border
        ws4.cell(row=row, column=3, value=fund.get("amount_range", "N/A")).border = border
        ws4.cell(row=row, column=4, value=fund.get("eligibility", "N/A")).border = border
    
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 30
    
    # Sheet 5: Economic Impact
    ws5 = wb.create_sheet("Economic Impact")
    ws5['A1'] = "Economic Impact Analysis"
    ws5['A1'].font = title_font
    
    impact = analysis.get("economic_impact", {})
    ws5['A3'] = "Scenario:"
    ws5['B3'] = impact.get("scenario", "N/A")
    
    formatted = impact.get("formatted", {})
    ws5['A5'] = "Property Damage:"
    ws5['B5'] = formatted.get("damage", "N/A")
    ws5['A6'] = "Business Interruption:"
    ws5['B6'] = formatted.get("interruption", "N/A")
    ws5['A7'] = "Recovery Time:"
    ws5['B7'] = formatted.get("recovery", "N/A")
    
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 35
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# API ENDPOINTS
# =============================================================================

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    """Main dashboard page."""
    if templates:
        return templates.TemplateResponse("dashboard.html", {
            "request": request,
            "user_types": [ut["name"] for ut in USER_TYPES]
        })
    else:
        # Fallback: return JSON API info
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Pythia - Mobile Bay Demo</title>
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
                       max-width: 900px; margin: 50px auto; padding: 20px; background: #f5f5f5; }}
                h1 {{ color: #2c5530; }}
                .card {{ background: white; padding: 20px; border-radius: 8px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                .endpoint {{ background: #e8f5e9; padding: 10px; margin: 5px 0; border-radius: 4px; font-family: monospace; }}
                .disclaimer {{ background: #fff3e0; padding: 15px; border-radius: 4px; border-left: 4px solid #ff9800; }}
                a {{ color: #1976d2; }}
            </style>
        </head>
        <body>
            <h1>🌊 Pythia - Mobile Bay Case Study Demo</h1>
            <p>Extreme Weather Risk Analysis for Private Equity Investors</p>
            
            <div class="disclaimer">
                <strong>Decision Support Tool</strong><br>
                {DISCLAIMER.replace(chr(10), '<br>')}
            </div>
            
            <div class="card">
                <h2>📍 Primary Prototype: {PRIMARY_PROTOTYPE['name']}</h2>
                <p>{PRIMARY_PROTOTYPE['description']}</p>
                <p>Coordinates: {PRIMARY_PROTOTYPE['coordinates']['lat']}, {PRIMARY_PROTOTYPE['coordinates']['lon']}</p>
            </div>
            
            <div class="card">
                <h2>🔗 API Endpoints</h2>
                <div class="endpoint">GET <a href="/api/demo">/api/demo</a> - Quick demo data</div>
                <div class="endpoint">GET <a href="/api/analyze/mobile-bay">/api/analyze/mobile-bay</a> - Full Mobile Bay analysis</div>
                <div class="endpoint">GET <a href="/api/risk-profile">/api/risk-profile</a> - Gulf Coast risk profile</div>
                <div class="endpoint">GET <a href="/api/solutions">/api/solutions</a> - Nature-based solutions</div>
                <div class="endpoint">GET <a href="/api/case-studies">/api/case-studies</a> - Mobile Bay case studies</div>
                <div class="endpoint">GET <a href="/api/funding">/api/funding</a> - Funding sources</div>
                <div class="endpoint">GET <a href="/api/economic-impact">/api/economic-impact</a> - Economic impact analysis</div>
                <div class="endpoint">GET <a href="/api/historical-events">/api/historical-events</a> - Historical weather events</div>
                <div class="endpoint">GET <a href="/api/openet">/api/openet</a> - OpenET API access</div>
                <div class="endpoint">GET <a href="/api/{api_name}">/api/{api_name}</a> - Generic API access</div>
                <div class="endpoint">GET <a href="/docs">/docs</a> - Interactive API documentation</div>
            </div>
            
            <div class="card">
                <h2>📊 Data Summary</h2>
                <p>Visit <a href="/api/data-summary">/api/data-summary</a> to see available data.</p>
            </div>
        </body>
        </html>
        """)


@app.get("/api/demo")
async def demo_data():
    """Quick demo endpoint showing system status and data availability."""
    summary = data_loader.get_all_data_summary()
    gulf_profile = get_gulf_coast_profile()
    
    return {
        "status": "operational",
        "system": "Pythia - Mobile Bay Case Study Demo",
        "timestamp": datetime.now().isoformat(),
        "primary_prototype": PRIMARY_PROTOTYPE,
        "disclaimer": DISCLAIMER,
        "data_summary": {
            "nature_based_solutions": summary.get("nature_based_solutions", {}).get("count", 0),
            "historical_events": summary.get("historical_events", {}).get("count", 0),
            "regional_profiles": summary.get("regions", {}).get("count", 0),
            "gulf_coast_primary_risks": list(gulf_profile.get("primary_risks", {}).keys())
        },
        "endpoints": [
            "/api/analyze/mobile-bay",
            "/api/risk-profile",
            "/api/solutions",
            "/api/case-studies",
            "/api/funding",
            "/api/economic-impact",
            "/api/historical-events",
            "/api/export/{format}"
        ]
    }


@app.get("/api/analyze/mobile-bay")
async def analyze_mobile_bay(
    investment_horizon: int = Query(default=5, description="Investment horizon in years (1-10)"),
    hurricane_scenario: int = Query(default=3, description="Hurricane category for scenario (1-5)")
):
    """
    Complete Mobile Bay analysis using REAL DATA.
    
    This is the main analysis endpoint that combines:
    - Gulf Coast regional risk profile
    - Nature-based solutions with local case studies
    - Economic impact analysis
    - Funding source recommendations
    """
    # Get real data
    gulf_profile = get_gulf_coast_profile()
    risk_assessment = calculate_risk_score(gulf_profile)
    solutions = get_mobile_bay_solutions()[:10]  # Top 10 solutions
    economic_impact = calculate_economic_impact(hurricane_scenario)
    funding = get_gulf_coast_funding()[:6]  # Top 6 funding sources
    historical = get_historical_gulf_events()
    
    # Get solutions with local case studies
    solutions_with_cases = [s for s in solutions if s.get("has_local_case_study")]
    
    return {
        "analysis_id": f"MB-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
        "timestamp": datetime.now().isoformat(),
        "location": PRIMARY_PROTOTYPE,
        "user_type": "Private Equity Investor",
        "investment_horizon_years": investment_horizon,
        "disclaimer": DISCLAIMER,
        
        "risk_assessment": {
            "composite_score": risk_assessment["composite_score"],
            "risk_level": risk_assessment["risk_level"],
            "confidence": risk_assessment["confidence"],
            "primary_risks": risk_assessment["breakdown"][:5],
            "data_source": "regional_risk_profiles.json - Gulf Coast profile"
        },
        
        "economic_impact": {
            "scenario": f"Category {hurricane_scenario} Hurricane",
            "potential_damage": economic_impact["formatted"]["damage"],
            "business_interruption": economic_impact["formatted"]["interruption"],
            "recovery_timeline": economic_impact["formatted"]["recovery"],
            "data_source": "economic_impact_data.json"
        },
        
        "recommended_solutions": [
            {
                "id": s["id"],
                "name": s["name"],
                "description": s["description"],
                "risk_types_addressed": s.get("risk_types", []),
                "has_local_case_study": s.get("has_local_case_study", False),
                "local_case_studies": s.get("mobile_bay_case_studies", []),
                "effectiveness": s.get("effectiveness_metrics", {}),
                "scale": s.get("scale", "unknown")
            }
            for s in solutions[:5]
        ],
        
        "local_case_studies": [
            {
                "solution": s["name"],
                "cases": s.get("mobile_bay_case_studies", [])
            }
            for s in solutions_with_cases[:3]
        ],
        
        "funding_sources": [
            {
                "name": f["name"],
                "category": f.get("category", ""),
                "eligible_uses": f.get("eligible_uses", [])[:3],
                "learn_more": f.get("learn_more", "")
            }
            for f in funding
        ],
        
        "historical_context": {
            "major_events": [
                {
                    "name": e["name"],
                    "date": e.get("date", ""),
                    "damage_cost": e.get("impact", {}).get("damage_cost", 0)
                }
                for e in historical[:3]
            ],
            "data_source": "historical_weather_events.json"
        },
        
        "export_available": True,
        "export_formats": ["json", "pdf", "excel"]
    }


@app.get("/api/risk-profile")
async def get_risk_profile(region: str = Query(default="gulf_coast")):
    """Get regional risk profile with calculated scores."""
    profiles = data_loader.load_regional_risk_profiles()
    profile = profiles.get("regions", {}).get(region, {})
    
    if not profile:
        return {"error": f"Region '{region}' not found", "available_regions": list(profiles.get("regions", {}).keys())}
    
    risk_assessment = calculate_risk_score(profile)
    
    return {
        "region": region,
        "name": profile.get("name", region),
        "coordinates": profile.get("coordinates", {}),
        "risk_assessment": risk_assessment,
        "primary_risks": profile.get("primary_risks", {}),
        "secondary_risks": profile.get("secondary_risks", {}),
        "vulnerable_assets": profile.get("vulnerable_assets", []),
        "adaptation_priorities": profile.get("adaptation_priorities", []),
        "biodiversity_risks": profile.get("biodiversity_risks", {}),
        "data_source": "regional_risk_profiles.json"
    }


@app.get("/api/solutions")
async def get_solutions(
    risk_type: Optional[str] = Query(default=None, description="Filter by risk type (flooding, hurricanes, storm_surge, etc.)"),
    local_only: bool = Query(default=False, description="Only show solutions with Mobile Bay case studies")
):
    """Get nature-based solutions relevant to Mobile Bay."""
    solutions = get_mobile_bay_solutions()
    
    if risk_type:
        solutions = [s for s in solutions if risk_type in s.get("risk_types", [])]
    
    if local_only:
        solutions = [s for s in solutions if s.get("has_local_case_study")]
    
    return {
        "count": len(solutions),
        "filters_applied": {"risk_type": risk_type, "local_only": local_only},
        "solutions": [
            {
                "id": s["id"],
                "name": s["name"],
                "description": s["description"],
                "risk_types": s.get("risk_types", []),
                "suitable_locations": s.get("suitable_locations", []),
                "scale": s.get("scale", ""),
                "benefits": s.get("benefits", []),
                "has_local_case_study": s.get("has_local_case_study", False),
                "case_study_count": len(s.get("mobile_bay_case_studies", [])),
                "effectiveness_metrics": s.get("effectiveness_metrics", {})
            }
            for s in solutions
        ],
        "data_source": "nature_based_solutions.json"
    }


@app.get("/api/case-studies")
async def get_case_studies():
    """Get all Mobile Bay/Alabama case studies from nature-based solutions."""
    solutions = get_mobile_bay_solutions()
    
    all_case_studies = []
    for solution in solutions:
        for case in solution.get("mobile_bay_case_studies", []):
            all_case_studies.append({
                "solution_id": solution["id"],
                "solution_name": solution["name"],
                "case_study": case
            })
    
    return {
        "location": "Mobile Bay, Alabama",
        "count": len(all_case_studies),
        "case_studies": all_case_studies,
        "data_source": "nature_based_solutions.json - case_studies field"
    }


@app.get("/api/funding")
async def get_funding():
    """Get funding sources for Gulf Coast resilience projects."""
    funding = get_gulf_coast_funding()
    
    return {
        "region": "Gulf Coast",
        "count": len(funding),
        "funding_sources": funding,
        "data_source": "funding_sources_NSB.json"
    }


@app.get("/api/economic-impact")
async def get_economic_impact(
    hurricane_category: int = Query(default=3, ge=1, le=5, description="Hurricane category (1-5)")
):
    """Get economic impact estimates for hurricane scenarios."""
    impact = calculate_economic_impact(hurricane_category)
    flood_data = data_loader.load_economic_impact_data().get("economic_impacts", {}).get("flood_impacts", {})
    
    return {
        "scenario": f"Category {hurricane_category} Hurricane",
        "hurricane_impact": impact,
        "flood_impact": flood_data,
        "disclaimer": "ROI analysis framework - no guarantees. Statistically significant and measurable improvements only.",
        "data_source": "economic_impact_data.json"
    }


@app.get("/api/historical-events")
async def get_historical_events():
    """Get historical weather events for Gulf Coast."""
    events = get_historical_gulf_events()
    all_events = data_loader.load_historical_weather_events().get("events", [])
    
    return {
        "gulf_coast_events": events,
        "total_events_in_database": len(all_events),
        "data_source": "historical_weather_events.json"
    }


@app.get("/api/infrastructure")
async def get_infrastructure():
    """Get coastal infrastructure needs data."""
    infra = get_coastal_infrastructure_data()
    return {
        "data": infra,
        "data_source": "coastal_areas_infrastructure_needs.json"
    }


@app.get("/api/user-types")
async def get_user_types():
    """Get available user types."""
    return {
        "user_types": USER_TYPES,
        "primary_prototype": "private_equity"
    }


@app.get("/api/data-summary")
async def get_data_summary():
    """Get summary of all available data."""
    summary = data_loader.get_all_data_summary()
    
    # Add additional data sources
    additional = {
        "funding_sources": len(get_gulf_coast_funding()),
        "infrastructure_data": "coastal_areas_infrastructure_needs.json"
    }
    
    return {
        **summary,
        "additional_data": additional,
        "primary_prototype": PRIMARY_PROTOTYPE
    }


@app.get("/api/export/{format}")
async def export_analysis(
    format: str,
    analysis_id: Optional[str] = Query(default=None)
):
    """Export analysis results in specified format."""
    # Generate fresh analysis
    analysis = await analyze_mobile_bay()
    
    if format == "json":
        return JSONResponse(content=analysis)
    elif format == "pdf":
        try:
            pdf_buffer = generate_pdf_report(analysis)
            filename = f"pythia_mobile_bay_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except Exception as e:
            return {"error": f"PDF generation failed: {str(e)}", "format": "pdf"}
    elif format == "excel":
        try:
            excel_buffer = generate_excel_report(analysis)
            filename = f"pythia_mobile_bay_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return StreamingResponse(
                excel_buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except Exception as e:
            return {"error": f"Excel generation failed: {str(e)}", "format": "excel"}
    else:
        return {"error": f"Unsupported format: {format}", "supported_formats": ["json", "pdf", "excel"]}


@app.get("/api/openet", response_class=JSONResponse)
def api_openet(
    lon: float = Query(..., description="Longitude"),
    lat: float = Query(..., description="Latitude"),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    source: str = Query("ensemble", description="OpenET data source (default: ensemble)"),
    use_cache: bool = Query(True, description="Use cache (default: True)")
):
    """Get OpenET evapotranspiration data via DataSourceManager."""
    result = data_loader.call_api_source(
        "openet_api", lon, lat, start_date, end_date, source, use_cache=use_cache
    )
    return result

@app.get("/api/{api_name}", response_class=JSONResponse)
def api_generic_api(
    api_name: str,
    use_cache: bool = Query(True, description="Use cache (default: True)"),
    q: str = Query("", description="Query string or parameters (optional)")
):
    """Call any registered API source by name. For advanced users/devs."""
    # Example: /api/bls_api?q=some_query
    # For most APIs, you may need to POST or provide structured params
    try:
        # For demonstration, pass q as a single argument if present
        args = [q] if q else []
        result = data_loader.call_api_source(api_name, *args, use_cache=use_cache)
        return result
    except Exception as e:
        return {"status": "ERROR", "error": str(e)}


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {
        "status": "healthy",
        "service": "Pythia - Mobile Bay Case Study Demo",
        "timestamp": datetime.now().isoformat(),
        "data_loader": "connected"
    }


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("\n" + "="*60)
    print("🌊 PYTHIA - Mobile Bay Case Study Demo")
    print("="*60)
    print(f"\n📍 Primary Prototype: {PRIMARY_PROTOTYPE['name']}")
    print(f"   {PRIMARY_PROTOTYPE['description']}")
    
    print("\n📊 Loading data...")
    try:
        summary = data_loader.get_all_data_summary()
        solutions = get_mobile_bay_solutions()
        local_cases = [s for s in solutions if s.get("has_local_case_study")]
        
        print(f"   ✅ {summary.get('nature_based_solutions', {}).get('count', 0)} nature-based solutions")
        print(f"   ✅ {len(local_cases)} solutions with Mobile Bay case studies")
        print(f"   ✅ {summary.get('historical_events', {}).get('count', 0)} historical weather events")
        print(f"   ✅ {summary.get('regions', {}).get('count', 0)} regional risk profiles")
        print(f"   ✅ {len(get_gulf_coast_funding())} funding sources")
    except Exception as e:
        print(f"   ⚠️ Data loading warning: {e}")
    
    print("\n🚀 Starting web server...")
    print("\n" + "-"*60)
    print("🌐 Open your browser to: http://localhost:8000")
    print("📋 API documentation:    http://localhost:8000/docs")
    print("🔍 Quick demo:           http://localhost:8000/api/demo")
    print("📊 Full analysis:        http://localhost:8000/api/analyze/mobile-bay")
    print("-"*60)
    print("\n⚠️  DISCLAIMER: Decision support tool only. No ROI guarantees.")
    print("="*60 + "\n")
    
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
